"""
Тест одностороннего экспорта базы в Excel (.xlsx) — Этап 2.1.
Собираем файл, читаем обратно openpyxl и проверяем шапку и данные.
"""
import io

from openpyxl import load_workbook

from app.db.models.agency import Agency
from app.repositories import user_repo
from app.schemas.apartment import ApartmentCreate
from app.services import apartment_service, excel_export_service


def _setup(db):
    agency = Agency(name="A", status="active", timezone="Asia/Tashkent", default_currency="USD")
    db.add(agency)
    db.flush()
    owner = user_repo.create(db, telegram_id=1, role="agency_admin", agency_id=agency.id, is_owner=True)
    db.commit()
    return agency.id, owner.id


def test_build_xlsx_has_header_and_rows(db):
    aid, uid = _setup(db)
    apartment_service.create_apartment(
        db, aid, created_by=uid,
        payload=ApartmentCreate(name="Квартира центр", price=55000, currency="USD", rooms=3),
    )
    apartment_service.create_apartment(
        db, aid, created_by=uid, payload=ApartmentCreate(name="Дом", price=120000),
    )

    data = excel_export_service.build_xlsx(db, aid)
    assert data[:2] == b"PK"  # .xlsx — это zip-архив

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "Объекты"
    header = [c.value for c in ws[1]]
    assert "Наименование" in header and "Цена" in header

    name_col = header.index("Наименование") + 1
    names = {ws.cell(row=r, column=name_col).value for r in range(2, ws.max_row + 1)}
    assert {"Квартира центр", "Дом"} <= names
