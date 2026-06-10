"""
Импорт готовой базы клиента из файла (.xlsx / .csv) — Этап 1.

Сценарий:
  1) Клиент в своём Google Sheets / Excel жмёт «Скачать как .xlsx (или .csv)» и
     загружает файл в приложение (наш Google-скоуп drive.file не умеет читать
     чужие таблицы, поэтому работаем именно с файлом — это ещё и универсально).
  2) Мы читаем заголовки и строки. Колонки у клиента названы по-своему
     («Цена объекта», «Кол-во комнат», «Р-н»…), поэтому Gemini автоматически
     сопоставляет их с нашими полями. Если ИИ недоступен — эвристика по названиям.
  3) Возвращаем превью: колонки клиента, первые строки и предложенный маппинг.
     Клиент подтверждает/правит сопоставление.
  4) По подтверждённому маппингу создаём объекты в БД (нормализуя значения:
     числа, валюта, enum-поля). Частичные данные — норма.

Файл на сервере НЕ хранится: при подтверждении клиент присылает его повторно
вместе с маппингом (stateless — без серверного кэша).
"""
import csv
import io
import json
import logging
from typing import Dict, List, Optional, Tuple

import httpx
from fastapi import status as http_status

from app.config import settings
from app.core.errors import AppError
from app.schemas.apartment import ApartmentCreate
from app.services import apartment_service
from app.services.listing_import_service import (
    CURRENCIES,
    LAND_AREA_TYPES,
    OBJ_COND_VALUES,
    OBJ_TYPE_VALUES,
)
from sqlalchemy.orm import Session

logger = logging.getLogger("uvicorn.error")

# ── Наши поля-цели для импорта (код → человекочитаемое описание для ИИ) ──
TARGET_FIELDS: List[Tuple[str, str]] = [
    ("name", "Наименование/заголовок объекта"),
    ("type", "Тип объекта: квартира, дом, коммерция, участок, земля"),
    ("district", "Район"),
    ("address", "Адрес, улица, ориентир, ЖК"),
    ("rooms", "Количество комнат"),
    ("floor", "Этаж"),
    ("total_floors", "Этажность дома (всего этажей)"),
    ("area", "Площадь в м²"),
    ("land_area", "Площадь участка в сотках"),
    ("condition", "Состояние/ремонт"),
    ("furniture_appliances", "Мебель и техника"),
    ("price", "Цена"),
    ("currency", "Валюта (USD/UZS/EUR)"),
    ("owner_phone", "Телефон собственника"),
    ("description", "Описание, дополнительная информация"),
    ("comment", "Внутренний комментарий"),
    ("source_link", "Ссылка-источник"),
    ("status", "Статус: активен, задаток, продан"),
]
TARGET_CODES = [c for c, _ in TARGET_FIELDS]

_INT_FIELDS = {"rooms", "floor", "total_floors"}
_FLOAT_FIELDS = {"area", "land_area", "price"}

_MAX_ROWS = 5000           # столько строк максимум импортируем за раз
_MAX_PREVIEW = 5           # столько строк-образцов показываем в превью
_MAX_COLS = 60             # защита от «широких» файлов
_MAX_FILE_BYTES = 8 * 1024 * 1024

# Подписи статуса/мебели → коды (для нормализации значений из файла).
_STATUS_HINTS = [("актив", "active"), ("задаток", "deposit"), ("депозит", "deposit"),
                 ("прода", "sold"), ("sold", "sold"), ("active", "active")]
_FA_NONE_HINTS = ("без мебел", "без техник", "ничего", "нет мебел", "пуст")


# ── Чтение файла ─────────────────────────────────────────────────────
def parse_file(filename: str, content: bytes) -> Tuple[List[str], List[List[str]]]:
    """Прочитать .xlsx/.csv → (заголовки, строки-значения). Строки — список строк."""
    if not content:
        raise AppError("import_file_empty", http_status.HTTP_400_BAD_REQUEST)
    if len(content) > _MAX_FILE_BYTES:
        raise AppError("import_file_too_big", http_status.HTTP_400_BAD_REQUEST)

    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        rows = _parse_csv(content)
    elif name.endswith(".xlsx") or name.endswith(".xlsm"):
        rows = _parse_xlsx(content)
    else:
        # Пытаемся по содержимому: xlsx — это zip (PK\x03\x04).
        rows = _parse_xlsx(content) if content[:2] == b"PK" else _parse_csv(content)

    # Убираем полностью пустые строки. Ячейки из xlsx ещё «сырые» (число, дата,
    # None) — приводим к строке через _cell_str, иначе .strip() падает на числах.
    rows = [r for r in rows if any(_cell_str(c).strip() for c in r)]
    if not rows:
        raise AppError("import_file_empty", http_status.HTTP_400_BAD_REQUEST)

    header = [(_cell_str(c)).strip() for c in rows[0]][:_MAX_COLS]
    width = len(header)
    data: List[List[str]] = []
    for r in rows[1:]:
        vals = [(_cell_str(c)).strip() for c in r][:width]
        if len(vals) < width:
            vals += [""] * (width - len(vals))
        data.append(vals)
        if len(data) >= _MAX_ROWS:
            break
    return header, data


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _parse_xlsx(content: bytes) -> List[List]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        logger.error("openpyxl не установлен: %s", exc)
        raise AppError("import_xlsx_unsupported", http_status.HTTP_503_SERVICE_UNAVAILABLE) from exc
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows
    except AppError:
        raise
    except Exception as exc:  # noqa: BLE001
        logger.info("Импорт базы: не удалось прочитать xlsx: %s", exc)
        raise AppError("import_file_unreadable", http_status.HTTP_400_BAD_REQUEST) from exc


def _parse_csv(content: bytes) -> List[List[str]]:
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise AppError("import_file_unreadable", http_status.HTTP_400_BAD_REQUEST)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [list(r) for r in reader]


# ── Сопоставление колонок (Gemini + эвристика) ───────────────────────
def suggest_mapping(header: List[str], rows: List[List[str]]) -> Dict[str, Optional[int]]:
    """{наше_поле -> индекс колонки клиента | None}. Сначала ИИ, иначе эвристика."""
    mapping = _heuristic_mapping(header)
    if settings.gemini_api_key:
        try:
            ai = _ai_mapping(header, rows)
            # ИИ имеет приоритет; пустые ответы добираем эвристикой.
            for f in TARGET_CODES:
                if ai.get(f) is not None:
                    mapping[f] = ai[f]
        except Exception as exc:  # noqa: BLE001
            logger.info("Импорт базы: ИИ-маппинг не удался, эвристика: %s", exc)
    return mapping


_HEUR = {
    "name": ("наимен", "название", "заголов", "объект", "title", "name"),
    "type": ("тип", "вид", "категор", "type"),
    "district": ("район", "р-н", "district", "mahalla", "tuman"),
    "address": ("адрес", "улиц", "ориентир", "жк", "address", "manzil"),
    "rooms": ("комнат", "комн", "rooms", "xona"),
    "floor": ("этаж", "floor", "qavat"),
    "total_floors": ("этажность", "всего этаж", "этажей", "total"),
    "area": ("площад", "кв.м", "м2", "м²", "area", "maydon"),
    "land_area": ("сот", "участ", "land", "yer"),
    "condition": ("состоян", "ремонт", "condition", "holat"),
    "furniture_appliances": ("мебел", "техник", "furniture"),
    "price": ("цена", "стоим", "price", "narx"),
    "currency": ("валют", "currency", "valyuta"),
    "owner_phone": ("телефон", "тел.", "номер", "phone", "контакт", "telefon"),
    "description": ("описан", "примечан", "коммент к объяв", "description", "izoh"),
    "comment": ("комментар", "заметк", "comment"),
    "source_link": ("ссылк", "источник", "url", "link", "havola"),
    "status": ("статус", "состояние сделк", "status", "holati"),
}


def _heuristic_mapping(header: List[str]) -> Dict[str, Optional[int]]:
    used: set = set()
    out: Dict[str, Optional[int]] = {f: None for f in TARGET_CODES}
    lowers = [h.lower().strip() for h in header]
    for field, hints in _HEUR.items():
        for idx, h in enumerate(lowers):
            if idx in used or not h:
                continue
            if any(hint in h for hint in hints):
                out[field] = idx
                used.add(idx)
                break
    return out


_GEMINI_URL = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"


def _ai_mapping(header: List[str], rows: List[List[str]]) -> Dict[str, Optional[int]]:
    # Колонки клиента: индекс, заголовок и пара примеров значений.
    cols_desc = []
    for i, h in enumerate(header):
        samples = [r[i] for r in rows[:3] if i < len(r) and (r[i] or "").strip()]
        cols_desc.append(f'{i}: "{h}" примеры: {samples}')
    fields_desc = "\n".join(f"- {c}: {d}" for c, d in TARGET_FIELDS)
    prompt = (
        "Есть таблица с объектами недвижимости. Сопоставь колонки таблицы с полями "
        "нашей системы. Верни ТОЛЬКО JSON-объект, где ключ — код нашего поля, "
        "значение — НОМЕР колонки таблицы (целое) или null, если подходящей нет.\n\n"
        "Наши поля:\n" + fields_desc + "\n\n"
        "Колонки таблицы (номер: \"заголовок\" примеры):\n" + "\n".join(cols_desc) + "\n\n"
        "Каждую колонку используй максимум один раз. Если сомневаешься — null."
    )
    url = _GEMINI_URL.format(model=settings.import_ai_model)
    payload = {
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0, "responseMimeType": "application/json"},
    }
    resp = httpx.post(url, params={"key": settings.gemini_api_key}, json=payload, timeout=45.0)
    resp.raise_for_status()
    data = resp.json()
    cands = data.get("candidates") or []
    parts = (cands[0].get("content") or {}).get("parts") or [] if cands else []
    content = "".join(p.get("text", "") for p in parts).strip() or "{}"
    raw = json.loads(content)

    width = len(header)
    out: Dict[str, Optional[int]] = {f: None for f in TARGET_CODES}
    used: set = set()
    for f in TARGET_CODES:
        v = raw.get(f)
        if isinstance(v, bool):  # bool — подкласс int, отсекаем
            continue
        if isinstance(v, int) and 0 <= v < width and v not in used:
            out[f] = v
            used.add(v)
    return out


# ── Нормализация значений ────────────────────────────────────────────
def _num(raw: str) -> Optional[float]:
    s = (raw or "").strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    s = "".join(ch for ch in s if ch.isdigit() or ch in ".-")
    if not s or s in (".", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _coerce(field: str, raw: str) -> object:
    raw = (raw or "").strip()
    if not raw:
        return None
    low = raw.lower()
    if field in _INT_FIELDS:
        v = _num(raw)
        return int(v) if v is not None and v >= 0 else None
    if field in _FLOAT_FIELDS:
        v = _num(raw)
        return v if v is not None and v >= 0 else None
    if field == "type":
        for t in OBJ_TYPE_VALUES:
            if t.lower() in low:
                return t
        return None
    if field == "condition":
        for c in OBJ_COND_VALUES:
            if c.lower() in low:
                return c
        return None
    if field == "currency":
        if any(k in low for k in ("usd", "$", "долл", "у.е", "y.e")):
            return "USD"
        if any(k in low for k in ("eur", "€", "евро")):
            return "EUR"
        if any(k in low for k in ("uzs", "сум", "сўм", "so'm", "som")):
            return "UZS"
        up = raw.upper()
        return up if up in CURRENCIES else None
    if field == "furniture_appliances":
        if any(k in low for k in _FA_NONE_HINTS):
            return "none"
        has_f = "мебел" in low or "furnitur" in low
        has_a = "техник" in low or "appliance" in low
        if has_f and has_a:
            return "furniture_and_appliances"
        if has_f:
            return "furniture_only"
        if has_a:
            return "appliances_only"
        return None
    if field == "status":
        for hint, code in _STATUS_HINTS:
            if hint in low:
                return code
        return None
    # name/district/address/owner_phone/description/comment/source_link — как есть.
    return raw


def _build_payload(mapping: Dict[str, Optional[int]], row: List[str]) -> dict:
    body: dict = {}
    for field in TARGET_CODES:
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            continue
        val = _coerce(field, row[idx])
        if val is not None:
            body[field] = val
    # Согласованность: дом/участок/земля — без «Этажа» (но «Этажность» остаётся);
    # квартира/коммерция — без «Соток».
    if body.get("type") in LAND_AREA_TYPES:
        body.pop("floor", None)
    else:
        body.pop("land_area", None)
    return body


# ── Высокоуровневые операции ─────────────────────────────────────────
def analyze(filename: str, content: bytes) -> dict:
    """Превью: колонки клиента, первые строки, предложенный маппинг."""
    header, rows = parse_file(filename, content)
    mapping = suggest_mapping(header, rows)
    return {
        "columns": header,
        "sample_rows": rows[:_MAX_PREVIEW],
        "total_rows": len(rows),
        "suggested_mapping": mapping,
        "target_fields": [{"code": c, "label": d} for c, d in TARGET_FIELDS],
    }


def commit(
    db: Session, agency_id: int, created_by: Optional[int],
    filename: str, content: bytes, mapping: Dict[str, Optional[int]],
) -> dict:
    """Создать объекты по подтверждённому маппингу. Вернуть статистику."""
    # Берём только известные поля и валидные индексы.
    clean_map: Dict[str, Optional[int]] = {}
    for f in TARGET_CODES:
        v = mapping.get(f)
        clean_map[f] = v if isinstance(v, int) and v >= 0 else None
    if all(v is None for v in clean_map.values()):
        raise AppError("import_no_mapping", http_status.HTTP_400_BAD_REQUEST)

    header, rows = parse_file(filename, content)
    created = skipped = failed = 0
    for row in rows:
        body = _build_payload(clean_map, row)
        if not body:
            skipped += 1
            continue
        try:
            apartment_service.create_apartment(
                db, agency_id, created_by=created_by, payload=ApartmentCreate(**body)
            )
            created += 1
        except AppError:
            skipped += 1
        except Exception as exc:  # noqa: BLE001
            logger.info("Импорт базы: строка не создана: %s", exc)
            failed += 1
    return {"created": created, "skipped": skipped, "failed": failed, "total": len(rows)}
