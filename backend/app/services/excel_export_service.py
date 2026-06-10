"""
Односторонний экспорт базы объектов в файл Excel (.xlsx) — Этап 2.1.

В отличие от Google Sheets (живая подключённая таблица), здесь просто отдаём
файл на скачивание: пользователь жмёт «Скачать Excel» и получает .xlsx со всеми
объектами агентства. Колонки и значения — те же, что и при выгрузке в Sheets
(переиспользуем sheets_service.export_matrix), поэтому формат единообразный.
"""
import io
import logging

from fastapi import status as http_status
from sqlalchemy.orm import Session

from app.core.errors import AppError
from app.services import sheets_service

logger = logging.getLogger("uvicorn.error")

# Ширины колонок в Sheets заданы в пикселях; в Excel — в «символах».
# Приблизительный перевод: ~7 пикселей на символ.
_PX_PER_CHAR = 7
# Колонки, которые в Sheets спрятаны как служебные (ID, «Изменено») — в Excel
# тоже прячем, чтобы файл выглядел чисто.
_HIDDEN_FIELDS = {"id", "updated_at"}


def build_xlsx(db: Session, agency_id: int) -> bytes:
    """Собрать .xlsx со всеми объектами агентства. Вернуть байты файла."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # noqa: BLE001
        logger.error("openpyxl недоступен для экспорта Excel: %s", exc)
        raise AppError("excel_unsupported", http_status.HTTP_503_SERVICE_UNAVAILABLE) from exc

    cols, values = sheets_service.export_matrix(db, agency_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты"

    for row in values:
        ws.append(row)

    # Жирная шапка + закрепление первой строки.
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"

    # Ширины, перенос/обрезка длинного текста, скрытие служебных колонок.
    wrap = Alignment(vertical="center", wrap_text=False)
    for idx, c in enumerate(cols):
        letter = get_column_letter(idx + 1)
        dim = ws.column_dimensions[letter]
        w = c.get("w")
        if w:
            dim.width = max(8, round(w / _PX_PER_CHAR))
        if c.get("f") in _HIDDEN_FIELDS:
            dim.hidden = True
    # Выравнивание значений по вертикали (без растягивания строк).
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
